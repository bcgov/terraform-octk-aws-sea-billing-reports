import logging
import os
import re
from itertools import groupby

import numpy as np
import pandas as pd
import requests
from jinja2 import Environment, FileSystemLoader

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table

logger = logging.getLogger(__name__)

grouping_columns = ['year', 'month', 'line_item_usage_account_id', 'Account_Name', 'Project', 'License_Plate',
					'Environment', 'Billing_Group', 'Owner_Name', 'Owner_Email', 'line_item_product_code', 'product_product_name']


def read_file_into_dataframe(local_file, accounts):
	conver_dict = {'line_item_usage_account_id': str}
	pd.set_option('display.float_format', '${:.2f}'.format)
	df = pd.read_csv(local_file, dtype=conver_dict)

	enhance_with_metadata(df, accounts)

	return df


def get_exchange_rate():
	usd_to_cad_rate = 1  # default/dummy value

	usd_to_cad_rate_param = os.getenv("FX_RATE")

	if usd_to_cad_rate_param:
		usd_to_cad_rate = float(usd_to_cad_rate_param)
	else:
		logger.info(f"No value provided for 'FX_RATE'. Will attempt to retrieve current rate via API service.")
		fx_api_key = os.getenv("FX_API_KEY")

		if not fx_api_key:
			logger.warning(f"No value provided for 'FX_RATE' or 'FX_API_KEY'. Please set a value via an environment variable. No conversion will be done.")
		else:
			parameters = {
				"q": "USD_CAD",
				"apiKey": fx_api_key,
				"compact": "ultra"
			}
			exchange_rate_response = requests.get("https://free.currconv.com/api/v7/convert", params=parameters)
			rate_response_json = exchange_rate_response.json()
			usd_to_cad_rate = rate_response_json["USD_CAD"]

	return usd_to_cad_rate


def enhance_with_metadata(df, accounts):
	account_details_by_account_id = make_account_by_id_lookup(accounts)

	core_billing_group = {
		"billing_group": "SEA Core",
		"admin_contact_email": "julian.subda@gov.bc.ca",
		"admin_contact_name": "Julian Subda",
		"Project": "Landing Zone Core",
		"Environment": "Core"
	}

	def get_account_name_element(account_id, element_index):
		account_email = account_details_by_account_id[account_id]['email']
		account_name = account_details_by_account_id[account_id]['name']

		if re.search("app", account_email):
			return account_name.split("-")[element_index]
		else:
			return account_name

	df['Billing_Group'] = df['line_item_usage_account_id'].apply(
		lambda x: account_details_by_account_id[x].get('billing_group',
													   core_billing_group['billing_group']))
	df['Owner_Name'] = df['line_item_usage_account_id'].apply(
		lambda x: account_details_by_account_id[x].get('admin_contact_name',
													   core_billing_group['admin_contact_name']))
	df['Owner_Email'] = df['line_item_usage_account_id'].apply(
		lambda x: account_details_by_account_id[x].get('admin_contact_email',
													   core_billing_group['admin_contact_email']))
	df['Project'] = df['line_item_usage_account_id'].apply(
		lambda x: account_details_by_account_id[x].get('Project',
													   core_billing_group['Project']))
	df['Environment'] = df['line_item_usage_account_id'].apply(
		lambda x: account_details_by_account_id[x].get('Environment',
													   core_billing_group['Environment']))
	df['Account_Name'] = df['line_item_usage_account_id'].apply(
		lambda x: account_details_by_account_id[x]['name'])
	df['License_Plate'] = df['line_item_usage_account_id'].apply(
		lambda x: get_account_name_element(x, 0))

	exchange_rate = get_exchange_rate()

	df['CAD'] = df['line_item_blended_cost'].apply(lambda x: x * exchange_rate)

def make_account_by_id_lookup(accounts):
	team_details_by_account_id = {}

	for account in accounts:
		team_details_by_account_id[account['id']] = account

	return team_details_by_account_id


def report(query_results_file, report_output_path, accounts, query_parameters):
	month = query_parameters['month']
	year = query_parameters['year']

	df = read_file_into_dataframe(query_results_file, accounts)

	core_billing_group = {
		"billing_group": "SEA Core",
		"admin_contact_email": "julian.subda@gov.bc.ca",
		"admin_contact_name": "Julian Subda",
		"name": "NA-NA"
	}

	sorted_accounts = sorted(accounts, key=lambda x: x.get('billing_group', core_billing_group['billing_group']))
	accounts_by_billing_group = groupby(sorted_accounts,
										key=lambda x: x.get('billing_group', core_billing_group['billing_group']))

	for billing_group, bg_accounts in accounts_by_billing_group:
		account_ids = [account["id"] for account in bg_accounts]

		billing_temp = df.query(
			f'year == [{year}] and month == [{month}] and (line_item_usage_account_id in {account_ids})')

		billing = pd.pivot_table(billing_temp,
								 index=grouping_columns,
								 values=['line_item_blended_cost', 'CAD'], aggfunc=[np.sum], fill_value=0, margins=True,
								 margins_name='Total')

		env = Environment(loader=FileSystemLoader('.'))
		template = env.get_template("report.html")

		template_vars = {
			"title": "AWS Report",
			"pivot_table": billing.to_html(),
			"business_unit": billing_group
		}

		html_out = template.render(template_vars)

		report_name = f"{year}-{month}-{billing_group}.html"
		report_file_name = f"{report_output_path}/{report_name}"

		with open(report_file_name, "w") as text_file:
			text_file.write(html_out)


def aggregate(query_results_file, query_parameters, summary_output_file):
	df = read_file_into_dataframe(query_results_file, query_parameters)

	df = df.groupby(grouping_columns).sum().reset_index()

	wb = Workbook()
	ws = wb.active

	for r in dataframe_to_rows(df, index=True, header=True):
		ws.append(r)

	ws.delete_rows(2)
	ws.delete_cols(1)

	# todo refactor to remove use of "magic" column letters (should find column by name instead)
	blended_cost_column = ws['M']
	cad_cost_col = ws['N']

	for cell in blended_cost_column:
		cell.number_format = '$#,##0.00'

	for cell in cad_cost_col:
		cell.number_format = '$#,##0.00'

	last_row = ws.max_row
	last_column = get_column_letter(ws.max_column)

	range_end = f"{last_column}{last_row}"

	table = Table(displayName="Charges", ref=f"A1:{range_end}")
	ws.add_table(table)

	column_widths = []
	for row in ws.rows:
		for i, cell in enumerate(row):
			if len(column_widths) > i:
				if len(str(cell.value)) > column_widths[i]:
					column_widths[i] = len(str(cell.value))
			else:
				column_widths += [len(str(cell.value))]

	for i, column_width in enumerate(column_widths):
		ws.column_dimensions[get_column_letter(i+1)].width = column_width + 4

	# todo refactor to remove use of "magic" column letters (should find column by name instead)
	ws.freeze_panes = ws['F2']

	wb.save(f"{summary_output_file}")
