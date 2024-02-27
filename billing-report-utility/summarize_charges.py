import logging
import os
import boto3
from datetime import date
import numpy as np
import pandas as pd
import requests
from jinja2 import Environment, FileSystemLoader
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table
import json
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

logger = logging.getLogger(__name__)

grouping_columns = [
    "year",
    "month",
    "line_item_usage_account_id",
    "Account_Name",
    "Project",
    "License_Plate",
    "Environment",
    "Billing_Group",
    "Owner_Name",
    "Owner_Email",
    "line_item_product_code",
    "product_product_name",
]

if os.environ.get("GROUP_TYPE") == "account_coding":
    grouping_columns.append("Account_Coding")


def read_file_into_dataframe(local_file, accounts):
    conver_dict = {"line_item_usage_account_id": str}
    pd.set_option("display.float_format", "${:.2f}".format)
    df = pd.read_csv(local_file, dtype=conver_dict)

    enhance_with_metadata(df, accounts)

    return df


def get_exchange_rate():
    # usd_to_cad_rate= 1
    AWS_REGION = "ca-central-1"
    ssm_client = boto3.client("ssm", region_name=AWS_REGION)
    url = "https://www.bankofcanada.ca/valet/observations/FXUSDCAD?recent=1"

    rc_channel = ssm_client.get_parameter(Name='/bcgov/billingutility/rocketchat_alert_webhook', WithDecryption=True)
    rc_channel_url = rc_channel['Parameter']['Value']
    teams_channel = ssm_client.get_parameter(Name='/bcgov/billingutility/teams_alert_webhook', WithDecryption=True)
    teams_channel_url = teams_channel['Parameter']['Value']

    requests_session = requests.Session()
    retries = Retry(
        total=3,
        backoff_factor=1,
        status_forcelist=[429, 500, 502, 503, 504, 404]
    )
    requests_session.mount("https://", HTTPAdapter(max_retries=retries))

    try:
        logger.info(
            f"requesting conversion rate from{url}"
        )
        response = requests_session.get(url)
        # print(response.text) # process response
    except Exception as error:
        print(error)
        attachment = [
            {
                "title": "Billing report utility failed",
                "text": f"The error occured is {error}",
                "color": "#764FA5",
            }
        ]
        rocketChatMessage = {
            "text": "Billing report utility failed with error",
            "attachments": attachment,
        }
        teamsMessage = {
            "type": "message",
            "attachments": [
                {
                    "contentType": "application/vnd.microsoft.card.adaptive",
                    "contentUrl": None,
                    "content": {
                        "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
                        "type": "AdaptiveCard",
                        "version": "1.2",
                        "body": [
                            {
                                "type": "TextBlock",
                                "wrap": True,
                                "text": f"The Billing utility failed due to {error}",
                            }
                        ]
                    }
                }
            ]
        }
        requests.post(
            rc_channel_url,
            data=json.dumps(rocketChatMessage),
            headers={
                "Content-Type": "application/json"},
        )
        requests.post(
            teams_channel_url,
            data=json.dumps(teamsMessage),
            headers={
                "Content-Type": "application/json"},
        )
    else:
        data = response.json()
        usd_to_cad_rate = float(data['observations'][0]['FXUSDCAD']['v'])
        # print(usd_to_cad_rate)
    return usd_to_cad_rate


def enhance_with_metadata(df, accounts):
    account_details_by_account_id = make_account_by_id_lookup(accounts)

    exchange_rate = get_exchange_rate()

    def get_account_metadata(account_id, field):
        account_details = account_details_by_account_id.get(account_id)
        if account_details:
            return account_details.get(field)
        else:
            return f"Missing Account: {account_id}"

    df["Account_Coding"] = df["line_item_usage_account_id"].apply(
        lambda x: get_account_metadata(x, "account_coding")
    )
    df["Billing_Group"] = df["line_item_usage_account_id"].apply(
        lambda x: get_account_metadata(x, "billing_group")
    )
    df["Owner_Name"] = df["line_item_usage_account_id"].apply(
        lambda x: get_account_metadata(x, "admin_contact_name")
    )
    df["Owner_Email"] = df["line_item_usage_account_id"].apply(
        lambda x: get_account_metadata(x, "admin_contact_email")
    )
    df["Additional_Contacts"] = df["line_item_usage_account_id"].apply(
    lambda x: get_account_metadata(x, "additional_contacts")
    )
    df["Project"] = df["line_item_usage_account_id"].apply(
        lambda x: get_account_metadata(x, "Project")
    )
    df["Environment"] = df["line_item_usage_account_id"].apply(
        lambda x: get_account_metadata(x, "Environment")
    )
    df["Account_Name"] = df["line_item_usage_account_id"].apply(
        lambda x: get_account_metadata(x, "name")
    )
    df["License_Plate"] = df["line_item_usage_account_id"].apply(
        lambda x: get_account_metadata(x, "license_plate")
    )
    df["CAD"] = df["line_item_blended_cost"].apply(lambda x: x * exchange_rate)


def make_account_by_id_lookup(accounts):
    team_details_by_account_id = {}

    for account in accounts:
        team_details_by_account_id[account["id"]] = account

    return team_details_by_account_id


def report(
    query_results_file,
    report_output_path,
    accounts,
    query_parameters,
    cb,
    quarterly_report_config,
):
    # Data frame relates account charges with account tags/ metadata which makes for easy aggregation
    df = read_file_into_dataframe(query_results_file, accounts)

    if os.environ.get("GROUP_TYPE") == "account_coding":
        billing_groups = set([account["account_coding"] for account in accounts])
    else:
        billing_groups = set([account["billing_group"] for account in accounts])

    # Total CAD for each billing group
    billing_group_totals = {}

    for billing_group in billing_groups:
        group_type = "Account_Coding" if os.environ.get("GROUP_TYPE") == "account_coding" else "Billing_Group"
        group_df = df.query(f'({group_type} == "{billing_group}")')

        sum_all_columns = group_df.sum(axis=0, skipna=True, numeric_only=True)
        sum_cad = sum_all_columns["CAD"]
        billing_group_totals[billing_group] = round(sum_cad, 2)

        billing = pd.pivot_table(
            group_df,
            index=grouping_columns,
            values=["line_item_blended_cost", "CAD"],
            aggfunc=[np.sum],
            fill_value=0,
            margins=True,
            margins_name="Total",
        )

        # todo pre-render
        env = Environment(loader=FileSystemLoader("."))
        template = env.get_template("templates/report.html.jinja2")

        template_vars = {
            "title": "Cloud Pathfinder Tenant Team Cloud Service Consumption Report (AWS)",
            "pivot_table": billing.to_html(),
            "business_unit": billing_group,
            "start_date": query_parameters["start_date"],
            "end_date": query_parameters["end_date"],
        }

        html_out = template.render(template_vars)

        format_string = "%Y-%m-%d"
        report_name = f"{date.strftime(query_parameters['start_date'], format_string)}-{date.strftime(query_parameters['end_date'], format_string)}-{billing_group}.html"
        report_file_name = f"{report_output_path}/{report_name}"

        with open(report_file_name, "w") as text_file:
            text_file.write(html_out)

        # invoke callback to pass back generated file to caller for current billing group
        cb(billing_group, report_file_name)

    if quarterly_report_config:
        format_string = "%Y-%m-%d"
        report_file_name = f"{report_output_path}/quarterly_report-{date.strftime(query_parameters['start_date'], format_string)}-{date.strftime(query_parameters['end_date'], format_string)}.xlsx"
        create_quarterly_excel(billing_group_totals, accounts, report_file_name)

        # invoke callback to pass back generated file to caller for current billing group
        cb("QUARTERLY_REPORT", report_file_name)

    return billing_group_totals


def aggregate(query_results_file, summary_output_path, accounts, query_parameters, cb):
    df = read_file_into_dataframe(query_results_file, accounts)

    format_string = "%Y-%m-%d"
    filename_prefix = f"{date.strftime(query_parameters['start_date'], format_string)}-{date.strftime(query_parameters['end_date'], format_string)}"

    create_excel(df, f"{summary_output_path}/charges-{filename_prefix}-ALL.xlsx")

    group_type = "account_coding" if os.environ.get("GROUP_TYPE") == "account_coding" else "billing_group"
    billing_groups = set([account[group_type] for account in accounts])

    for billing_group in billing_groups:
        group_type = "Account_Coding" if os.environ.get("GROUP_TYPE") == "account_coding" else "Billing_Group"
        group_df = df.query(f'({group_type} == "{billing_group}")')

        excel_output_path = (
            f"{summary_output_path}/charges-{filename_prefix}-{billing_group}.xlsx"
        )
        create_excel(group_df, excel_output_path)
        logger.debug(f"Done with  billing group '{billing_group}'...")
        # invoke callback to pass back generated file to caller for current billing group
        cb(billing_group, excel_output_path)


def set_to_formatted_string(some_set):
    formatted_list = ""
    for item in some_set:
        formatted_list = formatted_list + item + '; ' 
    return formatted_list.rstrip('; ')

def create_quarterly_excel(billing_group_totals, accounts, quarterly_output_file):
    wb = Workbook()
    ws = wb.active

    ws["A1"] = "Billing Group"
    ws["B1"] = "Total Spend (CAD)"
    ws["C1"] = "PO Names"
    ws["D1"] = "PO Emails"

    # NOTE: in this case billing_group = account_coding because for quarterly reports GROUP_TYPE = account_coding
    for billing_group, total in billing_group_totals.items():

        # Look up all the Product Owners for all the accounts with this account coding
        related_accounts = []

        for account in accounts :
          if account["account_coding"] == billing_group :
              related_accounts.append(account)

        # Create a duplicate-free list of all PO's associated with these accounts
        po_names = set([account["admin_contact_name"] for account in related_accounts])
        po_emails = set([account["admin_contact_email"] for account in related_accounts])

        po_names_formatted = set_to_formatted_string(po_names)
        po_emails_formatted = set_to_formatted_string(po_emails)
        
        if total != 0 :
          if billing_group == "000000000000000000000000":
              row = ("CPF. Pay direct via Service Order. No JV needed.", total, po_names_formatted, po_emails_formatted)
          else:           
              row = (billing_group, total, po_names_formatted, po_emails_formatted)
          ws.append(row)

    wb.save(f"{quarterly_output_file}")


def create_excel(df, summary_output_file):
    df = df.groupby(grouping_columns).sum().reset_index()

    wb = Workbook()
    ws = wb.active

    for r in dataframe_to_rows(df, index=True, header=True):
        ws.append(r)

    ws.delete_rows(2)
    ws.delete_cols(1)

    # todo refactor to remove use of "magic" column letters (should find column by name instead)
    blended_cost_column = ws["M"]
    cad_cost_col = ws["N"]

    for cell in blended_cost_column:
        cell.number_format = "$#,##0.00"

    for cell in cad_cost_col:
        cell.number_format = "$#,##0.00"

    last_row = ws.max_row
    last_column = get_column_letter(ws.max_column)
    range_end = f"{last_column}{last_row}"

    table = Table(displayName="Charges", ref=f"A1:{range_end}")
    ws.add_table(table)

    column_widths = []
    for row in ws.rows:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(str(cell.value)) > column_widths[i]:
                    column_widths[i] = len(str(cell.value))
            else:
                column_widths += [len(str(cell.value))]
    for i, column_width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = column_width + 4
    # todo refactor to remove use of "magic" column letters (should find column by name instead)
    ws.freeze_panes = ws["F2"]
    wb.save(f"{summary_output_file}")
